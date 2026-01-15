
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Output file path
output_file = r"e:\trae\0820焦甜香\数字化研发业务平台开发估算表_v6_250万版.xlsx"

# Data Definition
# Module, Function, Description, Frontend Days, Backend Days, AI/Algo Days
# Standard Rates: 2000 RMB/Day
# REVISION: Adjusted for ~2.5M Total Target (Reductions in AI Complexity)

dev_data = [
    # 5.1 Dashboard (System Home)
    ("1. 系统首页", "工作台配置", "拖拽式自定义布局、卡片管理", 10, 5, 0),
    ("1. 系统首页", "智能聚合", "跨系统待办集成、消息中心中心", 8, 12, 0),
    ("1. 系统首页", "全局搜索", "全域知识图谱索引、多模态检索", 10, 15, 20),
    ("1. 系统首页", "KPI驾驶舱", "多维指标可视化、实时计算引擎", 12, 15, 5),

    # 5.2 Market Insight
    ("2. 市场洞察", "舆情监测系统", "全网爬虫、NLP情感/意图识别大模型", 15, 20, 35), # Reduced from 40
    ("2. 市场洞察", "问卷调查平台", "可视化问卷设计、投放逻辑、回收统计", 12, 15, 0),
    ("2. 市场洞察", "竞品数据库", "多维竞品档案、自动对标分析", 10, 12, 8),

    # 5.3 Material Management
    ("3. 原料管理", "基地数字化", "GIS地图集成、IoT环境数据采集", 12, 15, 5),
    ("3. 原料管理", "原料基因库", "全生命周期档案、LIMS指纹图谱同步", 10, 20, 10),
    ("3. 原料管理", "智能库存", "效期预警、呆滞分析模型", 8, 12, 5),
    ("3. 原料管理", "采购辅助决策", "基于生产计划的需求预测算法", 8, 15, 15), # Reduced from 20

    # 5.4 Formula Management (CORE - HIGH COMPLEXITY)
    ("4. 配方管理", "配方精细化维护", "复杂结构编辑、版本回溯、实时成本", 20, 25, 5),
    ("4. 配方管理", "智能辅助设计", "基于多目标遗传算法(NSGA-II)的自动寻优", 15, 25, 45), # Reduced from 60
    ("4. 配方管理", "三级评审流程", "多级审批流、差异智能比对", 10, 15, 0),

    # 5.5 Flavor Management (CORE - HIGH COMPLEXITY)
    ("5. 香精香料", "香料知识库", "单体/香基档案、数字化香韵词典", 10, 15, 0),
    ("5. 香精香料", "AI智能调香", "感官-理化映射模型、竞品仿香逆向工程", 15, 20, 40), # Reduced from 50
    ("5. 香精香料", "精密配方管理", "高精度防差错集成、核心配方非对称加密", 15, 20, 10),

    # 5.6 Auxiliary Material
    ("6. 烟用辅材", "辅材适性匹配", "卷烟纸透气度与焦油量预测回归模型", 10, 15, 20), # Reduced from 25
    ("6. 烟用辅材", "在线包装设计", "WebGL 3D渲染引擎、在线批注协同", 25, 20, 5),
    ("6. 烟用辅材", "供应商协同", "资质自动审核、材料评价体系", 8, 12, 0),

    # 5.7 Process Management
    ("7. 工艺管理", "结构化工艺标准", "参数建模、版本严控、变更追溯", 15, 20, 0),
    ("7. 工艺管理", "工艺参数仿真", "CFD流体仿真集成、参数模拟预测", 15, 20, 35), # Reduced from 45
    ("7. 工艺管理", "MES交互接口", "机台状态实时校验、参数下发指令集", 8, 25, 0),

    # 5.8 LIMS
    ("8. 实验室管理", "全流程检测业务", "委托-采样-检测-报告全链路工作流", 20, 30, 0),
    ("8. 实验室管理", "仪器数据集成", "主流仪器(HPLC/GC)IoT接口、数据解析", 10, 25, 5),
    ("8. 实验室管理", "CNAS合规", "电子记录痕迹、质量体系审计追踪", 10, 15, 0),

    # 5.9 Quality Management
    ("9. 质量管理", "移动质检", "Pad端离线应用、多端数据同步", 20, 15, 0),
    ("9. 质量管理", "缺陷视觉识别", "外观缺陷CNN图像识别模型训练", 10, 15, 25), # Reduced from 30
    ("9. 质量管理", "QMS质量分析", "SPC统计过程控制、全链条质量追溯", 12, 18, 10),

    # 5.10 Online Experiment
    ("10. 在线试验", "高级排程系统", "基于资源约束的APS高级排程算法", 15, 20, 20), # Reduced from 25
    ("10. 在线试验", "过程实时监控", "TimeSeries高频时序数据采集(100Hz)", 15, 25, 10),
    ("10. 在线试验", "试验报告引擎", "多源异构数据自动聚合生成", 10, 15, 0),

    # 5.11 Comprehensive Mgmt
    ("11. 综合管理", "科研项目管理", "WBS分解、进度与经费关联监控", 12, 18, 5),
    ("11. 综合管理", "资产效能分析", "设备OEE分析、预测性维护(PdM)模型", 10, 15, 15),

    # 5.12 Research Assistant (Innovation)
    ("12. 科研助手", "科研知识图谱", "文献解析、百万级实体抽取、RAG构建", 10, 15, 40), # Reduced from 45
    ("12. 科研助手", "生成式AI应用", "智能摘要、多语种翻译、辅助写作Agent", 10, 15, 25), # Reduced from 30
]

# Create Dev DataFrame
df_dev = pd.DataFrame(dev_data, columns=["模块", "功能点", "描述", "前端人天", "后端人天", "AI/算法人天"])
df_dev["小计人天"] = df_dev["前端人天"] + df_dev["后端人天"] + df_dev["AI/算法人天"]

total_dev_days = df_dev["小计人天"].sum()

# Project Services (Overhead)
# Adjusted Ratios for Enterprise Delivery
overhead_data = [
    ("项目管理 (PM)", "全生命周期", "进度、质量、风险、沟通管理", 0.15),
    ("需求分析与架构设计", "前置阶段", "业务调研、微服务架构设计、领域建模", 0.20),
    ("系统测试与UAT", "质量保障", "单元/集成/性能/安全测试、用户验收", 0.20),
    ("部署实施与培训", "交付阶段", "多环境搭建、数据迁移、操作培训", 0.15),
]

services_rows = []
for name, phase, desc, ratio in overhead_data:
    days = int(total_dev_days * ratio)
    services_rows.append([name, phase, desc, 0, 0, 0, days])

df_services = pd.DataFrame(services_rows, columns=["模块", "功能点", "描述", "前端人天", "后端人天", "AI/算法人天", "小计人天"])

# Combine for final view
df_final = pd.concat([df_dev, df_services], ignore_index=True)

# Calculate Costs
DAILY_RATE = 2000
df_final["预估成本 (元)"] = df_final["小计人天"] * DAILY_RATE

total_days = df_final["小计人天"].sum()
total_cost = df_final["预估成本 (元)"].sum()

print(f"Total Man-Days: {total_days}")
print(f"Total Cost: ¥{total_cost:,.2f}")

# --- Generate Excel ---
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_final.to_excel(writer, index=False, sheet_name='详细估算表', startrow=2)
    
    workbook = writer.book
    worksheet = writer.sheets['详细估算表']
    
    # Title
    worksheet.merge_cells('A1:H1')
    title_cell = worksheet['A1']
    title_cell.value = "数字化研发业务平台 - 详细开发预算表 (精算版)"
    title_cell.font = Font(size=18, bold=True, color="333333")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Subtitle / Info
    worksheet.merge_cells('A2:H2')
    info_cell = worksheet['A2']
    info_cell.value = f"报价基准: {DAILY_RATE}元/人天 | 总工期预估: {int(total_days/22)} 人月"
    info_cell.alignment = Alignment(horizontal='right')
    info_cell.font = Font(italic=True, color="666666")

    # Header Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in range(1, 9):
        cell = worksheet.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Data Style
    for row in range(4, len(df_final) + 4):
        # Check if it's a Service row (at the end)
        is_service = row > (len(df_dev) + 3)
        
        for col in range(1, 9):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            
            if is_service and col <= 3:
                cell.font = Font(bold=True, color="2F5597")
            
            # Subtotal and Cost in Bold
            if col == 7 or col == 8:
                cell.font = Font(bold=True)
            
            # Cost formatting
            if col == 8:
                cell.number_format = '¥#,##0'
                
            # Gray out empty cells for Services
            if is_service and (col >= 4 and col <= 6):
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.value = "-"
                cell.alignment = Alignment(horizontal='center')

    # Total Row
    last_row = len(df_final) + 4
    worksheet.merge_cells(f'A{last_row}:F{last_row}')
    total_label = worksheet.cell(row=last_row, column=1)
    total_label.value = "项目总计"
    total_label.alignment = Alignment(horizontal='right')
    total_label.font = Font(size=14, bold=True)
    
    grand_days = worksheet.cell(row=last_row, column=7)
    grand_days.value = total_days
    grand_days.font = Font(size=14, bold=True, color="C00000")
    
    grand_cost = worksheet.cell(row=last_row, column=8)
    grand_cost.value = total_cost
    grand_cost.font = Font(size=14, bold=True, color="C00000")
    grand_cost.number_format = '¥#,##0'
    
    # Column Widths
    from openpyxl.utils import get_column_letter
    widths = [20, 25, 45, 10, 10, 12, 12, 18]
    for i, w in enumerate(widths):
        worksheet.column_dimensions[get_column_letter(i+1)].width = w

print(f"Enhanced Excel generated: {output_file}")

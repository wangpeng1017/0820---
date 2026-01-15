
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Output file path
output_file = r"e:\trae\0820焦甜香\数字化研发业务平台开发估算表_v5_260万版.xlsx"

# Data Definition
# Module, Function, Description, Frontend Days, Backend Days, AI/Algo Days
# Standard Rates: 2000 RMB/Day
# TARGET: ~2.6 Million RMB (~1300 Days total)
# STRATEGY: Increase AI/Algo and Backend complexity for Core Modules

dev_data = [
    # 5.1 Dashboard (System Home)
    ("1. 系统首页", "工作台配置", "拖拽式自定义布局、卡片管理", 6, 4, 0), # +2
    ("1. 系统首页", "智能聚合", "跨系统待办集成、消息中心中心", 5, 8, 0), # +3
    ("1. 系统首页", "全局搜索", "全域检索、关联推荐", 6, 8, 5), # +2 +2 +2
    ("1. 系统首页", "KPI驾驶舱", "多维指标可视化、实时计算引擎", 8, 10, 0), # +2 +2

    # 5.2 Market Insight (High Value)
    ("2. 市场洞察", "舆情监测系统", "全网爬虫、NLP情感识别", 10, 15, 8), # Reduced AI from 12 to 8 (-4)
    ("2. 市场洞察", "问卷调查平台", "可视化问卷设计、投放逻辑", 8, 10, 0),
    ("2. 市场洞察", "竞品数据库", "多维竞品档案、对标分析", 6, 8, 4), # Reduced AI from 5 to 4 (-1)

    # 5.3 Material Management
    ("3. 原料管理", "基地数字化", "GIS地图集成、IoT环境采集", 8, 10, 0), # +2 +2
    ("3. 原料管理", "原料基因库", "全生命周期档案、LIMS同步", 8, 12, 0), # +2 +2
    ("3. 原料管理", "智能库存", "效期预警、呆滞分析", 6, 8, 5), # +1 +2 +3
    ("3. 原料管理", "采购辅助决策", "需求预测算法", 6, 10, 8), # +1 +2 +4

    # 5.4 Formula Management (CORE)
    ("4. 配方管理", "配方精细化维护", "结构编辑、版本回溯、成本", 15, 20, 0),
    ("4. 配方管理", "智能辅助设计", "自动寻优、替代推荐算法", 12, 18, 10), # Reduced AI from 15 to 10 (-5)
    ("4. 配方管理", "三级评审流程", "多级审批流、差异比对", 6, 10, 0),

    # 5.5 Flavor Management (CORE)
    ("5. 香精香料", "香料知识库", "单体/香基档案、标准词典", 8, 10, 0), # +2 +2
    ("5. 香精香料", "AI智能调香", "仿香逆向、推荐模型", 12, 18, 15), # +2 +3 +5
    ("5. 香精香料", "精密配方管理", "防差错集成、加密存储", 10, 15, 0), # +2 +3

    # 5.6 Auxiliary Material
    ("6. 烟用辅材", "辅材适性匹配", "透气度与焦油量预测", 6, 8, 8), # +1 +2 +3
    ("6. 烟用辅材", "在线包装设计", "3D渲染预览、批注协同", 15, 12, 0), # +3 +2
    ("6. 烟用辅材", "供应商协同", "资质审核、评价体系", 5, 8, 0), # +1 +2

    # 5.7 Process Management
    ("7. 工艺管理", "结构化工艺标准", "参数建模、版本控制", 10, 15, 0), # +2 +3
    ("7. 工艺管理", "工艺参数仿真", "参数模拟预测模型", 10, 12, 12), # +2 +2 +4
    ("7. 工艺管理", "MES交互接口", "机台校验、参数下发", 6, 15, 0), # +1 +3

    # 5.8 LIMS
    ("8. 实验室管理", "全流程检测业务", "委托-检测-报告工作流", 12, 18, 0), # +2 +3
    ("8. 实验室管理", "仪器数据集成", "仪器IoT接口、数据解析", 6, 12, 0), # +1 +2
    ("8. 实验室管理", "CNAS合规", "电子记录、审计追踪", 6, 10, 0), # +1 +2

    # 5.9 Quality Management
    ("9. 质量管理", "移动质检", "Pad端应用、数据同步", 12, 8, 0), # +2 +2
    ("9. 质量管理", "缺陷视觉识别", "外观缺陷图像识别", 6, 8, 12), # +1 +2 +4
    ("9. 质量管理", "QMS质量分析", "SPC控制图、质量追溯", 8, 12, 5), # +2 +2 +3

    # 5.10 Online Experiment
    ("10. 在线试验", "高级排程系统", "机台冲突排程算法", 10, 12, 8), # +2 +2 +3
    ("10. 在线试验", "过程实时监控", "时序数据采集", 10, 15, 0), # +2 +3
    ("10. 在线试验", "试验报告引擎", "多源数据聚合", 6, 10, 0), # +1 +2

    # 5.11 Comprehensive Mgmt
    ("11. 综合管理", "科研项目管理", "WBS、进度经费监控", 8, 12, 0), # +2 +2
    ("11. 综合管理", "资产效能分析", "OEE分析", 6, 8, 0), # +1 +2

    # 5.12 Research Assistant
    ("12. 科研助手", "科研知识图谱", "文献解析、RAG构建", 8, 10, 15), # +2 +2 +5
    ("12. 科研助手", "生成式AI应用", "摘要、翻译、写作", 6, 8, 10), # +1 +2 +5
]

# Create Dev DataFrame
df_dev = pd.DataFrame(dev_data, columns=["模块", "功能点", "描述", "前端人天", "后端人天", "AI/算法人天"])
df_dev["小计人天"] = df_dev["前端人天"] + df_dev["后端人天"] + df_dev["AI/算法人天"]

total_dev_days = df_dev["小计人天"].sum()

# Project Services (Overhead)
# Target total days ~1300
overhead_data = [
    ("项目管理 (PM)", "全生命周期", "进度、质量、风险、沟通管理", 0.14),    # Reduced from 0.15 to 0.14
    ("需求分析与架构设计", "前置阶段", "业务调研、微服务架构设计、领域建模", 0.17), # Reduced from 0.18 to 0.17
    ("系统测试与UAT", "质量保障", "单元/集成/性能/安全测试、用户验收", 0.17), # Reduced from 0.18 to 0.17
    ("部署实施与培训", "交付阶段", "多环境搭建、数据迁移、操作培训", 0.10), # Kept 10%
]

services_rows = []
for name, phase, desc, ratio in overhead_data:
    days = int(total_dev_days * ratio)
    services_rows.append([name, phase, desc, 0, 0, 0, days])

df_services = pd.DataFrame(services_rows, columns=["模块", "功能点", "描述", "前端人天", "后端人天", "AI/算法人天", "小计人天"])

# Combine for final view
df_final = pd.concat([df_dev, df_services], ignore_index=True)

# Calculate Costs
DAILY_RATE = 2000
df_final["预估成本 (元)"] = df_final["小计人天"] * DAILY_RATE

total_days = df_final["小计人天"].sum()
total_cost = df_final["预估成本 (元)"].sum()

print(f"Total Man-Days: {total_days}")
print(f"Total Cost: {total_cost:,.2f} RMB")

# --- Generate Excel ---
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_final.to_excel(writer, index=False, sheet_name='详细估算表', startrow=2)
    
    workbook = writer.book
    worksheet = writer.sheets['详细估算表']
    
    # Title
    worksheet.merge_cells('A1:H1')
    title_cell = worksheet['A1']
    title_cell.value = "数字化研发业务平台 - 详细开发预算表 (260万版)"
    title_cell.font = Font(size=18, bold=True, color="333333")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Subtitle / Info
    worksheet.merge_cells('A2:H2')
    info_cell = worksheet['A2']
    info_cell.value = f"报价基准: {DAILY_RATE}元/人天 | 总工期预估: {int(total_days/22)} 人月"
    info_cell.alignment = Alignment(horizontal='right')
    info_cell.font = Font(italic=True, color="666666")

    # Header Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in range(1, 9):
        cell = worksheet.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Data Style
    for row in range(4, len(df_final) + 4):
        # Check if it's a Service row (at the end)
        is_service = row > (len(df_dev) + 3)
        
        for col in range(1, 9):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            
            if is_service and col <= 3:
                cell.font = Font(bold=True, color="2F5597")
            
            # Subtotal and Cost in Bold
            if col == 7 or col == 8:
                cell.font = Font(bold=True)
            
            # Cost formatting
            if col == 8:
                cell.number_format = '¥#,##0'
                
            # Gray out empty cells for Services
            if is_service and (col >= 4 and col <= 6):
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.value = "-"
                cell.alignment = Alignment(horizontal='center')

    # Total Row
    last_row = len(df_final) + 4
    worksheet.merge_cells(f'A{last_row}:F{last_row}')
    total_label = worksheet.cell(row=last_row, column=1)
    total_label.value = "项目总计"
    total_label.alignment = Alignment(horizontal='right')
    total_label.font = Font(size=14, bold=True)
    
    grand_days = worksheet.cell(row=last_row, column=7)
    grand_days.value = total_days
    grand_days.font = Font(size=14, bold=True, color="C00000")
    
    grand_cost = worksheet.cell(row=last_row, column=8)
    grand_cost.value = total_cost
    grand_cost.font = Font(size=14, bold=True, color="C00000")
    grand_cost.number_format = '¥#,##0'
    
    # Column Widths
    widths = [20, 25, 45, 10, 10, 12, 12, 18]
    for i, w in enumerate(widths):
        worksheet.column_dimensions[get_column_letter(i+1)].width = w

print(f"Enhanced Excel generated: {output_file}")
